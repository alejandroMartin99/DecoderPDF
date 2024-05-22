from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def save_format_frontend(output_path, dataframes, sheet_names):

    wb = Workbook()

    for df, sheet_name in zip(dataframes, sheet_names):
        if df.empty:
            continue

        ws = wb.create_sheet(title=sheet_name)

        column_width = 16
        for col_num, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = column_width

        for col_num, value in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=value).font = Font(bold=True)
            ws.cell(row=1, column=col_num).fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        for row_num, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

                if "_Check" in df.columns[col_num - 1]:
                    ws.column_dimensions[get_column_letter(col_num)].width = 5
                    if value == "X":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    elif value == "Y":
                        cell.fill = PatternFill(start_color="78DE78", end_color="78DE78", fill_type="solid")

        max_column_letter = get_column_letter(ws.max_column)
        ref = f"A1:{max_column_letter}{ws.max_row}"
        table = Table(displayName=sheet_name, ref=ref)

        style = TableStyleInfo(
            name="TableStyleLight1", 
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    wb.remove(wb.active)
    wb.save(output_path)


